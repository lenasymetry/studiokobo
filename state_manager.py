# Contenu de state_manager.py
# Gestion des callbacks et de l'état de session

import streamlit as st
import openpyxl
import json
import datetime
import copy

from project_definitions import (
    get_default_dims_19,
    get_default_door_props_19,
    get_default_drawer_props,
    get_default_drawer_props_19,
    get_default_joue_props,
    get_default_vertical_divider_props,
    get_default_vertical_shelf_props,
)
from machining_logic import calculate_all_zones_2d

def get_default_shelf_props():
    """Retourne les propriétés par défaut pour une étagère."""
    return {
        'height': 200.0,
        'thickness': 19.0,
        'shelf_type': 'mobile',
        'zone_id': None,
        'material': 'Matière Corps'
    }

def get_default_debit_data():
    """Retourne les données de débit par défaut avec les pièces de base."""
    return [
        {
            "Référence Pièce": "Traverse Bas",
            "Longueur (mm)": 0,
            "Largeur (mm)": 0,
            "Epaisseur": 0,
            "Chant Avant": True,
            "Chant Arrière": True,
            "Chant Gauche": False,
            "Chant Droit": False,
            "Usinage": ""
        },
        {
            "Référence Pièce": "Traverse Haut",
            "Longueur (mm)": 0,
            "Largeur (mm)": 0,
            "Epaisseur": 0,
            "Chant Avant": True,
            "Chant Arrière": True,
            "Chant Gauche": False,
            "Chant Droit": False,
            "Usinage": ""
        },
        {
            "Référence Pièce": "Montant Gauche",
            "Longueur (mm)": 0,
            "Largeur (mm)": 0,
            "Epaisseur": 0,
            "Chant Avant": True,
            "Chant Arrière": True,
            "Chant Gauche": True,
            "Chant Droit": True,
            "Usinage": ""
        },
        {
            "Référence Pièce": "Montant Droit",
            "Longueur (mm)": 0,
            "Largeur (mm)": 0,
            "Epaisseur": 0,
            "Chant Avant": True,
            "Chant Arrière": True,
            "Chant Gauche": True,
            "Chant Droit": True,
            "Usinage": ""
        },
        {
            "Référence Pièce": "Fond",
            "Longueur (mm)": 0,
            "Largeur (mm)": 0,
            "Epaisseur": 0,
            "Chant Avant": False,
            "Chant Arrière": False,
            "Chant Gauche": False,
            "Chant Droit": False,
            "Usinage": ""
        }
    ]

def get_selected_cabinet():
    idx = st.session_state.get('selected_cabinet_index')
    if idx is not None and idx < len(st.session_state['scene_cabinets']): return st.session_state['scene_cabinets'][idx]
    return None

def initialize_session_state():
    """Initialise l'état de session global."""
    st.session_state.setdefault('scene_cabinets', [])
    st.session_state.setdefault('selected_cabinet_index', None)
    st.session_state.setdefault('base_cabinet_index', 0)
    st.session_state.setdefault('unit_select', 'mm')

    # Infos Globales du Projet
    st.session_state.setdefault('project_name', "Nouveau Projet")
    st.session_state.setdefault('corps_meuble', "Caisson 1")
    st.session_state.setdefault('quantity', 1)
    st.session_state.setdefault('client', "CLIENT NOM")
    st.session_state.setdefault('adresse_chantier', "") # AJOUTÉ
    st.session_state.setdefault('ref_chantier', "")
    st.session_state.setdefault('telephone', "")
    st.session_state.setdefault('date_souhaitee', datetime.date.today())
    st.session_state.setdefault('panneau_decor', "BLANC")
    st.session_state.setdefault('chant_mm', "1mm")
    st.session_state.setdefault('decor_chant', "BLANC")
    
    # Propriétés des pieds
    st.session_state.setdefault('has_feet', False)
    st.session_state.setdefault('foot_height', 80.0) 
    st.session_state.setdefault('foot_diameter', 30.0)
    
    # Pose en 2 temps (prévisualisation -> validation)
    # pending_placement: { kind: 'shelf'|'vertical_divider'|'vertical_shelf', cabinet_index: int, props: dict }
    st.session_state.setdefault('pending_placement', None)

    # Exports : mémorise le JSON de scène au moment de la dernière génération
    # None = jamais généré, chaîne = scène déjà générée
    st.session_state.setdefault('exports_scene_json', None)
    # Résultats stockés des derniers exports générés
    st.session_state.setdefault('exports_html_data', None)
    st.session_state.setdefault('exports_html_ok', False)
    st.session_state.setdefault('exports_dwg_data', None)
    st.session_state.setdefault('exports_dwg_ok', False)
    st.session_state.setdefault('exports_dwg_filename', None)
    st.session_state.setdefault('exports_xls_data', None)
    st.session_state.setdefault('exports_material_filter_key', None)
    st.session_state.setdefault('exports_material_scope', "Toutes les matières")
    st.session_state.setdefault('exports_selected_materials', [])
    st.session_state.setdefault('exports_skp_data', None)
    st.session_state.setdefault('exports_skp_ok', False)
    st.session_state.setdefault('exports_skp_has_doors', False)
    st.session_state.setdefault('exports_skp_closed_data', None)
    st.session_state.setdefault('exports_skp_closed_ok', False)
    st.session_state.setdefault('exports_skp_open_data', None)
    st.session_state.setdefault('exports_skp_open_ok', False)

def load_save_state():
    if 'file_loader' in st.session_state and st.session_state.file_loader is not None:
        uploaded_file = st.session_state.file_loader
        try:
            workbook = openpyxl.load_workbook(uploaded_file)
            if 'SaveData' in workbook.sheetnames:
                save_sheet = workbook['SaveData']
                # Le JSON peut être découpé sur plusieurs lignes (limite Excel de
                # 32767 caractères par cellule) : on reassemble tous les morceaux.
                chunks = []
                for (cell,) in save_sheet.iter_rows(min_col=1, max_col=1, values_only=False):
                    if cell.value is None:
                        break
                    chunks.append(str(cell.value))
                json_data_str = "".join(chunks)
                if json_data_str:
                    loaded_data = json.loads(json_data_str)
                    st.session_state['project_name'] = loaded_data.get('project_name', 'Nouveau Projet')
                    st.session_state['client'] = loaded_data.get('client', '')
                    st.session_state['adresse_chantier'] = loaded_data.get('adresse_chantier', '') # AJOUTÉ
                    st.session_state['ref_chantier'] = loaded_data.get('ref_chantier', '')
                    st.session_state['telephone'] = loaded_data.get('telephone', '')
                    if 'date_souhaitee' in loaded_data:
                         st.session_state['date_souhaitee'] = datetime.date.fromisoformat(loaded_data['date_souhaitee'])
                    st.session_state['panneau_decor'] = loaded_data.get('panneau_decor', '')
                    st.session_state['chant_mm'] = loaded_data.get('chant_mm', '')
                    st.session_state['decor_chant'] = loaded_data.get('decor_chant', '')
                    st.session_state['has_feet'] = loaded_data.get('has_feet', False)
                    st.session_state['foot_height'] = loaded_data.get('foot_height', 80.0)
                    st.session_state['foot_diameter'] = loaded_data.get('foot_diameter', 50.0)
                    st.session_state['scene_cabinets'] = loaded_data.get('scene_cabinets', [])
                    if st.session_state['scene_cabinets']:
                        st.session_state['selected_cabinet_index'] = 0
                        st.session_state['base_cabinet_index'] = 0
                    else:
                        st.session_state['selected_cabinet_index'] = None
                        st.session_state['base_cabinet_index'] = 0
                    st.success("Projet chargé.")
                    st.rerun()
        except Exception as e:
            st.error(f"Erreur chargement : {e}")


def _to_float(value, default=0.0):
    try:
        return float(value)
    except Exception:
        return float(default)


def _to_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return bool(default)
    if isinstance(value, (int, float)):
        return value != 0
    txt = str(value).strip().lower()
    if txt in {'1', 'true', 'yes', 'on'}:
        return True
    if txt in {'0', 'false', 'no', 'off', 'none', ''}:
        return False
    return bool(value)


def _cm_key_to_mm(data, keys, default_mm=0.0):
    for key in keys:
        if key in data and data.get(key) is not None:
            return _to_float(data.get(key), default_mm / 10.0) * 10.0
    return float(default_mm)


def _norm_type(value: str) -> str:
    txt = str(value or '').strip().lower()
    txt = txt.replace('é', 'e').replace('è', 'e').replace('ê', 'e')
    txt = txt.replace('-', '_').replace(' ', '_')
    return txt


def _parse_element_key_geometry(key_value):
    """Décode une clé configurateur du type y0_x0_y1_x1 en cm.

    Retourne un dict avec bottom_cm/top_cm/x_left_cm/x_right_cm si la clé est
    exploitable, sinon None.
    """
    if key_value is None:
        return None
    raw = str(key_value).strip().replace(',', '.')
    parts = [p for p in raw.split('_') if p != '']
    if len(parts) != 4:
        return None
    try:
        y0_cm, x0_cm, y1_cm, x1_cm = [float(p) for p in parts]
    except Exception:
        return None

    bottom_cm = min(y0_cm, y1_cm)
    top_cm = max(y0_cm, y1_cm)
    x_left_cm = min(x0_cm, x1_cm)
    x_right_cm = max(x0_cm, x1_cm)
    return {
        'bottom_cm': bottom_cm,
        'top_cm': top_cm,
        'x_left_cm': x_left_cm,
        'x_right_cm': x_right_cm,
    }


def _infer_panel_thickness_mm(payload: dict, meuble: dict, default_mm=18.0):
    candidates = []

    def _add_from_elements(elements, section_width_cm):
        if not elements or section_width_cm <= 0:
            return
        for el in elements:
            etype = _norm_type(el.get('type', ''))
            if etype not in {'shelf', 'shelf_simple', 'shelf_fixed', 'etagere', 'etagere_simple', 'etagere_fixe'}:
                continue
            xl = _to_float(el.get('x_left_cm', 0.0), 0.0)
            xr = _to_float(el.get('x_right_cm', section_width_cm), section_width_cm)
            if abs(xl) > 0.2:
                continue
            thickness_cm = (section_width_cm - xr) / 2.0
            thickness_mm = thickness_cm * 10.0
            if 12.0 <= thickness_mm <= 30.0:
                candidates.append(thickness_mm)

    _add_from_elements(payload.get('elements', []) or [], _to_float(meuble.get('L', 0.0), 0.0))
    tv_side = payload.get('tvSide', {}) if isinstance(payload.get('tvSide', {}), dict) else {}
    for side_key in ('left', 'right', 'top'):
        side = tv_side.get(side_key)
        if isinstance(side, dict):
            _add_from_elements(side.get('elements', []) or [], _to_float(side.get('W_cm', 0.0), 0.0))

    if candidates:
        return round(sum(candidates) / len(candidates), 1)
    return float(default_mm)


def _add_box_spec(boxes, *, x_mm, y_mm, z_mm, w_mm, d_mm, h_mm, kind, color_group):
    if w_mm <= 0.0 or d_mm <= 0.0 or h_mm <= 0.0:
        return
    boxes.append({
        'x_mm': float(x_mm),
        'y_mm': float(y_mm),
        'z_mm': float(z_mm),
        'w_mm': float(w_mm),
        'd_mm': float(d_mm),
        'h_mm': float(h_mm),
        'kind': str(kind),
        'color_group': str(color_group),
    })


def _add_shell_geometry(boxes, *, x_mm, z_mm, L_mm, H_mm, P_mm, t_mm):
    inner_w = max(0.0, L_mm - 2.0 * t_mm)
    inner_h = max(0.0, H_mm - 2.0 * t_mm)
    _add_box_spec(boxes, x_mm=x_mm, y_mm=0.0, z_mm=z_mm, w_mm=t_mm, d_mm=P_mm, h_mm=H_mm, kind='upright_left', color_group='body')
    _add_box_spec(boxes, x_mm=x_mm + L_mm - t_mm, y_mm=0.0, z_mm=z_mm, w_mm=t_mm, d_mm=P_mm, h_mm=H_mm, kind='upright_right', color_group='body')
    _add_box_spec(boxes, x_mm=x_mm + t_mm, y_mm=0.0, z_mm=z_mm, w_mm=inner_w, d_mm=P_mm, h_mm=t_mm, kind='bottom', color_group='body')
    _add_box_spec(boxes, x_mm=x_mm + t_mm, y_mm=0.0, z_mm=z_mm + H_mm - t_mm, w_mm=inner_w, d_mm=P_mm, h_mm=t_mm, kind='top', color_group='body')
    _add_box_spec(boxes, x_mm=x_mm + t_mm, y_mm=max(0.0, P_mm - t_mm), z_mm=z_mm + t_mm, w_mm=inner_w, d_mm=t_mm, h_mm=inner_h, kind='back', color_group='body')


def _build_web_render_geometry(payload: dict, meuble: dict, t_mm: float):
    boxes = []
    meuble_L_mm = _cm_key_to_mm(meuble, ['L', 'L_cm'], 0.0)
    meuble_H_mm = _cm_key_to_mm(meuble, ['H', 'H_cm'], 0.0)
    meuble_P_mm = _cm_key_to_mm(meuble, ['P', 'P_cm'], 0.0)

    # Caisson bas principal
    _add_shell_geometry(boxes, x_mm=0.0, z_mm=0.0, L_mm=meuble_L_mm, H_mm=meuble_H_mm, P_mm=meuble_P_mm, t_mm=t_mm)

    # Elements du meuble bas
    for el in payload.get('elements', []) or []:
        etype = _norm_type(el.get('type', ''))
        key_geom = _parse_element_key_geometry(el.get('key'))
        if etype in {'separator', 'separateur'}:
            x_mm = _cm_key_to_mm(el, ['x_cm'], 0.0)
            z_mm = _cm_key_to_mm(el, ['y_cm'], t_mm)
            _add_box_spec(
                boxes,
                x_mm=x_mm,
                y_mm=0.0,
                z_mm=z_mm,
                w_mm=t_mm,
                d_mm=max(0.0, meuble_P_mm - t_mm),
                h_mm=max(0.0, meuble_H_mm - z_mm - t_mm),
                kind='separator',
                color_group='body',
            )
            continue

        if etype in {'shelf', 'shelf_simple', 'shelf_fixed', 'etagere', 'etagere_simple', 'etagere_fixe'}:
            y_mm = _cm_key_to_mm(el, ['y_cm'], t_mm)
            x_left_mm = _cm_key_to_mm(el, ['x_left_cm'], t_mm)
            x_right_mm = _cm_key_to_mm(el, ['x_right_cm'], meuble_L_mm - t_mm)
            _add_box_spec(
                boxes,
                x_mm=x_left_mm,
                y_mm=0.0,
                z_mm=max(0.0, y_mm - t_mm),
                w_mm=max(0.0, x_right_mm - x_left_mm),
                d_mm=meuble_P_mm,
                h_mm=t_mm,
                kind='shelf',
                color_group='body',
            )
            continue

        if etype in {'porte', 'door', 'door_single', 'porte_simple', 'single_door', 'door_double', 'porte_double', 'double_door'} and key_geom is not None:
            _add_box_spec(
                boxes,
                x_mm=key_geom['x_left_cm'] * 10.0,
                y_mm=-t_mm,
                z_mm=key_geom['bottom_cm'] * 10.0,
                w_mm=(key_geom['x_right_cm'] - key_geom['x_left_cm']) * 10.0,
                d_mm=t_mm,
                h_mm=(key_geom['top_cm'] - key_geom['bottom_cm']) * 10.0,
                kind='door',
                color_group='accessory',
            )
            continue

        if etype in {'tiroir', 'drawer', 'drawer_simple', 'tiroir_simple', 'drawer_front', 'tiroir_face'}:
            if key_geom is None:
                continue
            _add_box_spec(
                boxes,
                x_mm=key_geom['x_left_cm'] * 10.0,
                y_mm=-t_mm,
                z_mm=key_geom['bottom_cm'] * 10.0,
                w_mm=(key_geom['x_right_cm'] - key_geom['x_left_cm']) * 10.0,
                d_mm=t_mm,
                h_mm=(key_geom['top_cm'] - key_geom['bottom_cm']) * 10.0,
                kind='drawer_front',
                color_group='accessory',
            )

    return {
        'panel_thickness_mm': float(t_mm),
        'boxes': boxes,
        'total_dims_mm': {
            'L': meuble_L_mm,
            'H': meuble_H_mm,
            'P': meuble_P_mm,
        },
    }


def _build_side_cabinet(side_payload: dict, *, side_name: str, attachment_dir: str,
                        parent_index: int, depth_mm: float, panel_t_mm: float,
                        material_name: str):
    dims_default = get_default_dims_19()
    L_mm = _cm_key_to_mm(side_payload, ['W_cm', 'L_cm', 'largeur_cm'], 0.0)
    H_mm = _cm_key_to_mm(side_payload, ['H_cm', 'height_cm', 'hauteur_cm'], 0.0)
    if L_mm <= 0.0 or H_mm <= 0.0:
        return None

    dims = copy.deepcopy(dims_default)
    dims['L_raw'] = L_mm
    dims['H_raw'] = H_mm
    dims['W_raw'] = depth_mm
    dims['t_lr_raw'] = panel_t_mm
    dims['t_tb_raw'] = panel_t_mm
    dims['t_fb_raw'] = panel_t_mm

    cabinet = {
        'dims': dims,
        'debit_data': get_default_debit_data(),
        'name': side_name,
        'parent_index': parent_index,
        'attachment_dir': attachment_dir,
        'door_props': copy.deepcopy(get_default_door_props_19()),
        'drawer_props': get_default_drawer_props_19(),
        'drawers': [],
        'shelves': [],
        'material_body': material_name,
        'vertical_dividers': [],
        'vertical_shelves': [],
        'base_elements': {
            'has_back_panel': True,
            'has_left_upright': True,
            'has_right_upright': True,
            'has_bottom_traverse': True,
            'has_top_traverse': True,
        },
        'joues': {
            'gauche': get_default_joue_props(),
            'droite': get_default_joue_props(),
            'dessus': get_default_joue_props(),
            'dessous': get_default_joue_props(),
        },
    }

    t_tb = float(panel_t_mm)
    for el in side_payload.get('elements', []) or []:
        etype = _norm_type(el.get('type', ''))
        if etype not in {'shelf', 'shelf_simple', 'shelf_fixed', 'etagere', 'etagere_simple', 'etagere_fixe'}:
            continue
        y_mm = _cm_key_to_mm(el, ['y_cm', 'z_cm', 'height_cm'], -1.0)
        if y_mm < 0.0:
            key_geom = _parse_element_key_geometry(el.get('key'))
            if key_geom is not None:
                y_mm = key_geom['top_cm'] * 10.0
        if y_mm < 0.0:
            continue
        cabinet['shelves'].append({
            'height': max(0.0, min(H_mm - 2.0 * t_tb, y_mm - t_tb)),
            'thickness': panel_t_mm,
            'shelf_type': 'fixe',
            'zone_id': None,
            'material': material_name,
        })

    for shelf in cabinet['shelves']:
        stored_x_start = shelf.get('stored_shelf_x_start_mm')
        stored_width = shelf.get('stored_shelf_width_mm')
        shelf_bottom = float(shelf.get('height', 0.0)) + t_tb
        shelf_top = shelf_bottom + float(shelf.get('thickness', 19.0))
        if stored_x_start is None or stored_width is None or not zones_for_import:
            shelf['zone_id'] = None
            continue

        shelf_x_left = float(stored_x_start)
        shelf_x_right = shelf_x_left + float(stored_width)
        chosen_zone = None
        best_score = 0.0
        for zone in zones_for_import:
            zx_min = float(zone['x_min'])
            zx_max = float(zone['x_max'])
            x_match = abs(zx_min - shelf_x_left) <= tol_mm and abs(zx_max - shelf_x_right) <= tol_mm
            if x_match:
                chosen_zone = zone
                break
            overlap_x = max(0.0, min(shelf_x_right, zx_max) - max(shelf_x_left, zx_min))
            if overlap_x <= 0.0:
                continue
            score = overlap_x
            if score > best_score + 1e-6:
                best_score = score
                chosen_zone = zone
        if chosen_zone is not None:
            zone_x_min = float(chosen_zone['x_min'])
            zone_x_max = float(chosen_zone['x_max'])
            shelf['zone_id'] = chosen_zone['id']
            shelf['stored_zone_coords'] = {
                'x_min': chosen_zone['x_min'],
                'x_max': chosen_zone['x_max'],
                'y_min': chosen_zone['y_min'],
                'y_max': chosen_zone['y_max'],
            }
            shelf['stored_shelf_x_start_mm'] = max(shelf_x_left, zone_x_min)
            shelf['stored_shelf_width_mm'] = max(0.0, min(shelf_x_right, zone_x_max) - shelf['stored_shelf_x_start_mm'])
            if shelf['stored_shelf_width_mm'] <= 0.0:
                shelf['stored_shelf_x_start_mm'] = zone_x_min
                shelf['stored_shelf_width_mm'] = max(0.0, zone_x_max - zone_x_min)
        else:
            shelf['zone_id'] = None

    cabinet['shelves'] = sorted(cabinet['shelves'], key=lambda s: float(s.get('height', 0.0)))
    return cabinet


def _build_scene_from_web_config(payload: dict):
    meuble = payload.get('meuble', {}) if isinstance(payload.get('meuble', {}), dict) else {}
    central = _build_cabinet_from_web_config(payload)
    scene = [central]

    material_name = str(
        payload.get('mat')
        or meuble.get('matLabel')
        or meuble.get('mat')
        or 'Matière Corps'
    ).strip() or 'Matière Corps'
    depth_mm = float(central['dims']['W_raw'])
    panel_t_mm = float(central['dims']['t_tb_raw'])

    tv_side = payload.get('tvSide', {}) if isinstance(payload.get('tvSide', {}), dict) else {}
    for side_key, side_name, attachment_dir in (
        ('left', 'Caisson 1 (AS Gauche)', 'left'),
        ('right', 'Caisson 2 (AS Droite)', 'right'),
        ('top', 'Caisson 3 (AS Haut)', 'up'),
    ):
        side_payload = tv_side.get(side_key)
        if not isinstance(side_payload, dict):
            continue
        side_cabinet = _build_side_cabinet(
            side_payload,
            side_name=side_name,
            attachment_dir=attachment_dir,
            parent_index=0,
            depth_mm=depth_mm,
            panel_t_mm=panel_t_mm,
            material_name=material_name,
        )
        if side_cabinet is not None:
            scene.append(side_cabinet)

    return scene


def _build_cabinet_from_web_config(payload: dict):
    dims_default = get_default_dims_19()
    door_default = get_default_door_props_19()
    drawer_default = get_default_drawer_props()
    meuble = payload.get('meuble', {}) if isinstance(payload.get('meuble', {}), dict) else {}
    inferred_t_mm = _infer_panel_thickness_mm(payload, meuble, default_mm=18.0)

    L_mm = _cm_key_to_mm(payload, ['L_cm', 'largeur_cm', 'width_cm'], 0.0)
    H_mm = _cm_key_to_mm(payload, ['H_cm', 'hauteur_cm', 'height_cm'], 0.0)
    P_mm = _cm_key_to_mm(payload, ['P_cm', 'profondeur_cm', 'depth_cm'], 0.0)
    if L_mm <= 0.0:
        L_mm = _cm_key_to_mm(meuble, ['L', 'L_cm', 'largeur_cm', 'width_cm'], dims_default['L_raw'])
    if H_mm <= 0.0:
        H_mm = _cm_key_to_mm(meuble, ['H', 'H_cm', 'hauteur_cm', 'height_cm'], dims_default['H_raw'])
    if P_mm <= 0.0:
        P_mm = _cm_key_to_mm(meuble, ['P', 'P_cm', 'profondeur_cm', 'depth_cm'], dims_default['W_raw'])
    if L_mm <= 0.0:
        L_mm = float(dims_default['L_raw'])
    if H_mm <= 0.0:
        H_mm = float(dims_default['H_raw'])
    if P_mm <= 0.0:
        P_mm = float(dims_default['W_raw'])

    dims = copy.deepcopy(dims_default)
    dims['L_raw'] = L_mm
    dims['H_raw'] = H_mm
    dims['W_raw'] = P_mm
    dims['t_lr_raw'] = inferred_t_mm
    dims['t_tb_raw'] = inferred_t_mm
    dims['t_fb_raw'] = inferred_t_mm

    mat_name = str(
        payload.get('mat')
        or meuble.get('matLabel')
        or meuble.get('mat')
        or 'Matière Corps'
    ).strip()
    if not mat_name:
        mat_name = 'Matière Corps'

    cabinet = {
        'dims': dims,
        'debit_data': get_default_debit_data(),
        'name': f"Caisson 0 ({str(payload.get('furniture_type') or payload.get('type') or 'Import JSON').strip() or 'Import JSON'})",
        'parent_index': None,
        'attachment_dir': None,
        'door_props': copy.deepcopy(door_default),
        'imported_doors': [],
        'drawer_props': get_default_drawer_props_19(),
        'drawers': [],
        'shelves': [],
        'material_body': mat_name,
        'vertical_dividers': [],
        'vertical_shelves': [],
        'base_elements': {
            'has_back_panel': True,
            'has_left_upright': True,
            'has_right_upright': True,
            'has_bottom_traverse': True,
            'has_top_traverse': True,
        },
        'joues': {
            'gauche': get_default_joue_props(),
            'droite': get_default_joue_props(),
            'dessus': get_default_joue_props(),
            'dessous': get_default_joue_props(),
        },
        'web_import_payload': payload,
    }

    t_tb = float(dims.get('t_tb_raw', 19.0))
    t_lr = float(dims.get('t_lr_raw', 19.0))
    interior_x_max_mm = max(0.0, L_mm - 2.0 * t_lr)

    def _normalize_web_x_bounds(x_left_mm: float, x_right_mm: float):
        left = float(x_left_mm)
        right = float(x_right_mm)
        if right < left:
            left, right = right, left

        # Les coordonnees web sont generalement exprimees dans l'interieur utile du caisson.
        if left >= -2.0 and right <= interior_x_max_mm + 2.0:
            left += t_lr
            right += t_lr

        left = max(0.0, min(L_mm, left))
        right = max(left, min(L_mm, right))
        return left, right

    def _normalize_web_separator_left_x(x_left_mm: float):
        x_left = float(x_left_mm)
        if 0.0 <= x_left <= interior_x_max_mm + 2.0:
            x_left += t_lr
        x_left = max(t_lr, min(L_mm - t_lr - inferred_t_mm, x_left))
        return x_left

    def _normalize_web_bottom_y(y_bottom_mm: float):
        y_bottom = float(y_bottom_mm)
        interior_h_mm = max(0.0, H_mm - 2.0 * t_tb)
        if 0.0 <= y_bottom <= interior_h_mm + 2.0:
            y_bottom += t_tb
        return max(0.0, min(H_mm, y_bottom))

    def _parse_opening(value):
        if value is None:
            return None
        txt = _norm_type(value)
        if txt in {'left', 'gauche', 'ouverture_gauche', 'door_left'}:
            return 'left'
        if txt in {'right', 'droite', 'ouverture_droite', 'door_right'}:
            return 'right'
        return None

    def _parse_is_double_door(el: dict, etype: str):
        if etype in {'door_double', 'porte_double', 'double_door', '2_battants', 'deux_battants'}:
            return True
        door_type = _norm_type(el.get('door_type', el.get('porte_type', '')))
        if door_type in {'double', 'double_door', 'porte_double', '2_battants', 'deux_battants'}:
            return True
        battants_local_raw = el.get('battants', el.get('door_battants', None))
        if battants_local_raw is not None:
            battants_local = int(_to_float(battants_local_raw, 1))
            return battants_local >= 2
        return bool(global_door_is_double)

    elements = payload.get('elements', []) or []

    # Reglage porte depuis le JSON (top-level ou elements)
    door_props = copy.deepcopy(door_default)
    door_props['has_door'] = _to_bool(payload.get('has_door', False), False)
    battants = int(_to_float(payload.get('battants', payload.get('door_battants', 0)), 0))
    door_type_raw = _norm_type(payload.get('door_type', payload.get('porte_type', '')))
    if door_props['has_door']:
        if battants >= 2 or door_type_raw in {'double', 'double_door', 'porte_double', '2_battants', 'deux_battants'}:
            door_props['door_type'] = 'double'
        else:
            door_props['door_type'] = 'single'
    elif battants >= 2:
        door_props['has_door'] = True
        door_props['door_type'] = 'double'
    elif battants == 1 or door_type_raw in {'single', 'simple', 'single_door', 'porte_simple', '1_battant', 'un_battant'}:
        door_props['has_door'] = True
        door_props['door_type'] = 'single'

    global_door_is_double = (
        battants >= 2
        or door_type_raw in {'double', 'double_door', 'porte_double', '2_battants', 'deux_battants'}
    )
    global_opening_raw = payload.get('door_opening', payload.get('opening', payload.get('sens_ouverture', None)))

    # Type de tiroir par defaut unique pour tout le projet importe.
    drawer_system_raw = _norm_type(payload.get('drawer_system', 'tandembox'))
    if drawer_system_raw in {'legrabox', 'legra', 'blum_legrabox'}:
        drawer_system_default = 'LÉGRABOX'
        drawer_tech_default = 'K'
    elif drawer_system_raw in {'anglaise', 'tiroir_anglaise'}:
        drawer_system_default = 'ANGLAISE'
        drawer_tech_default = 'K'
    else:
        drawer_system_default = 'TANDEMBOX'
        drawer_tech_default = 'K'

    pending_drawers = []
    pending_doors = []
    for el in elements:
        etype = _norm_type(el.get('type', ''))
        key_geom = _parse_element_key_geometry(el.get('key'))

        if etype in {'separator', 'separateur'}:
            x_mm = _cm_key_to_mm(el, ['x_cm'], -1.0)
            if x_mm >= 0.0:
                sep_left_x = _normalize_web_separator_left_x(x_mm)
                sep_center_x = sep_left_x + inferred_t_mm / 2.0
                existing_divider = None
                for div in cabinet['vertical_dividers']:
                    if abs(float(div.get('position_x', 0.0)) - sep_center_x) <= 1.0:
                        existing_divider = div
                        break

                interior_h_mm = max(0.0, H_mm - 2.0 * t_tb)
                sep_bottom_y = 0.0
                sep_top_y = interior_h_mm
                force_full_height = False

                # Permet d'exprimer explicitement un montant pleine hauteur côté JSON.
                full_height_raw = el.get('full_height', el.get('is_full_height', None))
                has_full_height_flag = full_height_raw is not None
                if has_full_height_flag and _to_bool(full_height_raw, False):
                    force_full_height = True

                # Cas legacy web: separator avec seulement x_cm/y_cm = montant pleine hauteur.
                # La hauteur partielle doit être exprimée explicitement (key, y_bottom/y_top ou h_cm).
                has_explicit_vertical_bounds = (
                    key_geom is not None
                    or ('y_bottom_cm' in el and el.get('y_bottom_cm') is not None)
                    or ('bottom_cm' in el and el.get('bottom_cm') is not None)
                    or ('z_bottom_cm' in el and el.get('z_bottom_cm') is not None)
                    or ('y_top_cm' in el and el.get('y_top_cm') is not None)
                    or ('top_cm' in el and el.get('top_cm') is not None)
                    or ('z_top_cm' in el and el.get('z_top_cm') is not None)
                    or ('h_cm' in el and el.get('h_cm') is not None)
                )
                if not has_full_height_flag and not has_explicit_vertical_bounds:
                    force_full_height = True

                # Priorité 1: géométrie explicite via key (y0_x0_y1_x1) si présente.
                if key_geom is not None and not force_full_height:
                    sep_bottom_y = max(0.0, min(interior_h_mm, key_geom['bottom_cm'] * 10.0))
                    sep_top_y = max(sep_bottom_y, min(interior_h_mm, key_geom['top_cm'] * 10.0))
                else:
                    # Priorité 2: bornes verticales explicites bottom/top dans le JSON.
                    has_bottom = False
                    has_top = False
                    for k in ('y_bottom_cm', 'bottom_cm', 'z_bottom_cm'):
                        if k in el and el.get(k) is not None:
                            sep_bottom_y = _to_float(el.get(k), 0.0) * 10.0
                            has_bottom = True
                            break
                    for k in ('y_top_cm', 'top_cm', 'z_top_cm'):
                        if k in el and el.get(k) is not None:
                            sep_top_y = _to_float(el.get(k), interior_h_mm / 10.0) * 10.0
                            has_top = True
                            break

                    # Priorité 3: y_cm + h_cm explicites.
                    if not force_full_height and 'y_cm' in el and el.get('y_cm') is not None:
                        y_base = _to_float(el.get('y_cm'), 0.0) * 10.0
                        if 'h_cm' in el and el.get('h_cm') is not None:
                            h_val = max(0.0, _to_float(el.get('h_cm'), 0.0) * 10.0)
                            sep_bottom_y = y_base
                            sep_top_y = y_base + h_val
                            has_bottom = True
                            has_top = True
                        elif not has_bottom and not has_top:
                            # Compatibilité legacy: y_cm seul = départ vertical, sommet à l'intérieur haut.
                            sep_bottom_y = y_base
                            sep_top_y = interior_h_mm

                    if force_full_height:
                        sep_bottom_y = 0.0
                        sep_top_y = interior_h_mm

                sep_bottom_y = max(0.0, min(interior_h_mm, float(sep_bottom_y)))
                sep_top_y = max(sep_bottom_y, min(interior_h_mm, float(sep_top_y)))

                if existing_divider is not None:
                    existing_full = bool(existing_divider.get('_force_full_height', False))
                    if force_full_height or existing_full:
                        existing_divider['bottom_y'] = 0.0
                        existing_divider['top_y'] = interior_h_mm
                        existing_divider['_force_full_height'] = True
                    else:
                        existing_bottom = float(existing_divider.get('bottom_y', 0.0))
                        existing_top = float(existing_divider.get('top_y', interior_h_mm))
                        existing_divider['bottom_y'] = max(0.0, min(existing_bottom, sep_bottom_y))
                        existing_divider['top_y'] = min(interior_h_mm, max(existing_top, sep_top_y))
                        existing_divider['_force_full_height'] = False
                    # Conserver une épaisseur cohérente et la matière importée la plus récente.
                    existing_divider['thickness'] = max(float(existing_divider.get('thickness', inferred_t_mm)), float(inferred_t_mm))
                    existing_divider['material'] = mat_name
                    existing_divider['_imported_separator'] = True
                else:
                    cabinet['vertical_dividers'].append({
                        'position_x': sep_center_x,
                        'thickness': inferred_t_mm,
                        'material': mat_name,
                        'bottom_y': sep_bottom_y,
                        'top_y': sep_top_y,
                        '_force_full_height': bool(force_full_height),
                        '_imported_separator': True,
                    })
            continue

        if etype in {
            'shelf', 'shelf_simple', 'shelf_fixed', 'etagere', 'etagere_simple', 'etagere_fixe'
        }:
            y_mm = _cm_key_to_mm(el, ['y_cm', 'z_cm', 'height_cm'], -1.0)
            if y_mm < 0.0 and key_geom is not None:
                y_mm = key_geom['top_cm'] * 10.0
            if y_mm < 0.0:
                y_mm = t_tb
            x_left_mm = _cm_key_to_mm(el, ['x_left_cm'], 0.0)
            x_right_mm = _cm_key_to_mm(el, ['x_right_cm'], interior_x_max_mm)
            if key_geom is not None:
                x_left_mm = key_geom['x_left_cm'] * 10.0
                x_right_mm = key_geom['x_right_cm'] * 10.0
            shelf_x_start_mm, shelf_x_end_mm = _normalize_web_x_bounds(x_left_mm, x_right_mm)
            shelf_width_mm = max(0.0, shelf_x_end_mm - shelf_x_start_mm)
            # Le modele interne attend une position depuis le dessus de
            # traverse basse, puis bornee a l'interieur utile.
            shelf_height = max(0.0, min(H_mm - 2.0 * t_tb, y_mm - t_tb))
            shelf_data = {
                'height': shelf_height,
                'thickness': 19.0,
                'shelf_type': 'fixe',
                'zone_id': None,
                'material': mat_name,
            }
            if shelf_width_mm > 0.0:
                shelf_data['stored_shelf_x_start_mm'] = shelf_x_start_mm
                shelf_data['stored_shelf_width_mm'] = shelf_width_mm
            cabinet['shelves'].append(shelf_data)
            continue

        if etype in {
            'drawer', 'drawer_simple', 'tiroir', 'tiroir_simple', 'drawer_front', 'tiroir_face'
        }:
            y_mm = _cm_key_to_mm(el, ['y_cm', 'bottom_cm', 'z_cm'], -1.0)
            h_mm = _cm_key_to_mm(el, ['h_cm', 'height_cm', 'face_h_cm'], -1.0)
            x_left_mm = _cm_key_to_mm(el, ['x_left_cm'], 0.0)
            x_right_mm = _cm_key_to_mm(el, ['x_right_cm'], interior_x_max_mm)
            if key_geom is not None:
                if y_mm < 0.0:
                    y_mm = key_geom['bottom_cm'] * 10.0
                if h_mm < 0.0:
                    h_mm = (key_geom['top_cm'] - key_geom['bottom_cm']) * 10.0
                x_left_mm = key_geom['x_left_cm'] * 10.0
                x_right_mm = key_geom['x_right_cm'] * 10.0
            if y_mm < 0.0:
                y_mm = 0.0
            if h_mm <= 0.0:
                h_mm = 150.0

            x_left_mm, x_right_mm = _normalize_web_x_bounds(x_left_mm, x_right_mm)
            y_mm = _normalize_web_bottom_y(y_mm)
            pending_drawers.append({
                'x_left_mm': x_left_mm,
                'x_right_mm': x_right_mm,
                'y_bottom_mm': y_mm,
                'height_mm': h_mm,
            })
            continue

        if etype in {
            'door', 'door_single', 'porte', 'porte_simple', 'single_door', '1_battant', 'un_battant'
        }:
            door_props['has_door'] = True
            door_props['door_type'] = 'single'
            x_left_mm = _cm_key_to_mm(el, ['x_left_cm'], 0.0)
            x_right_mm = _cm_key_to_mm(el, ['x_right_cm'], interior_x_max_mm)
            y_mm = _cm_key_to_mm(el, ['y_cm', 'bottom_cm', 'z_cm'], -1.0)
            h_mm = _cm_key_to_mm(el, ['h_cm', 'height_cm', 'face_h_cm'], -1.0)
            if key_geom is not None:
                x_left_mm = key_geom['x_left_cm'] * 10.0
                x_right_mm = key_geom['x_right_cm'] * 10.0
                if y_mm < 0.0:
                    y_mm = key_geom['bottom_cm'] * 10.0
                if h_mm < 0.0:
                    h_mm = (key_geom['top_cm'] - key_geom['bottom_cm']) * 10.0
            if h_mm <= 0.0:
                h_mm = max(100.0, H_mm - 2.0 * t_tb)
            if y_mm < 0.0:
                y_mm = t_tb
            x_left_mm, x_right_mm = _normalize_web_x_bounds(x_left_mm, x_right_mm)
            y_mm = _normalize_web_bottom_y(y_mm)
            is_double_leaf = _parse_is_double_door(el, etype)
            if is_double_leaf:
                door_props['door_type'] = 'double'
            opening_raw = el.get('opening', el.get('door_opening', el.get('opening_side', el.get('sens_ouverture', global_opening_raw))))
            parsed_opening = _parse_opening(opening_raw)
            pending_doors.append({
                'x_left_mm': x_left_mm,
                'x_right_mm': x_right_mm,
                'y_bottom_mm': y_mm,
                'height_mm': h_mm,
                'door_opening': parsed_opening,
                'has_explicit_opening': parsed_opening in {'left', 'right'},
                'is_double_leaf': is_double_leaf,
            })
            continue

        if etype in {
            'door_double', 'porte_double', 'double_door', '2_battants', 'deux_battants'
        }:
            door_props['has_door'] = True
            door_props['door_type'] = 'double'
            x_left_mm = _cm_key_to_mm(el, ['x_left_cm'], 0.0)
            x_right_mm = _cm_key_to_mm(el, ['x_right_cm'], interior_x_max_mm)
            y_mm = _cm_key_to_mm(el, ['y_cm', 'bottom_cm', 'z_cm'], -1.0)
            h_mm = _cm_key_to_mm(el, ['h_cm', 'height_cm', 'face_h_cm'], -1.0)
            if key_geom is not None:
                x_left_mm = key_geom['x_left_cm'] * 10.0
                x_right_mm = key_geom['x_right_cm'] * 10.0
                if y_mm < 0.0:
                    y_mm = key_geom['bottom_cm'] * 10.0
                if h_mm < 0.0:
                    h_mm = (key_geom['top_cm'] - key_geom['bottom_cm']) * 10.0
            if h_mm <= 0.0:
                h_mm = max(100.0, H_mm - 2.0 * t_tb)
            if y_mm < 0.0:
                y_mm = t_tb
            x_left_mm, x_right_mm = _normalize_web_x_bounds(x_left_mm, x_right_mm)
            y_mm = _normalize_web_bottom_y(y_mm)
            opening_raw = el.get('opening', el.get('door_opening', el.get('opening_side', el.get('sens_ouverture', global_opening_raw))))
            parsed_opening = _parse_opening(opening_raw)
            pending_doors.append({
                'x_left_mm': x_left_mm,
                'x_right_mm': x_right_mm,
                'y_bottom_mm': y_mm,
                'height_mm': h_mm,
                'door_opening': parsed_opening,
                'has_explicit_opening': parsed_opening in {'left', 'right'},
                'is_double_leaf': True,
            })
            continue

    # Filet de sécurité import web:
    # si des séparateurs sont absents/incomplets, reconstruire les montants
    # à partir des bornes X des éléments (key, x_left_cm, x_right_cm).
    interior_h_mm = max(0.0, H_mm - 2.0 * t_tb)
    boundary_candidates_left_x = []
    for el in elements:
        key_geom = _parse_element_key_geometry(el.get('key'))
        x_values_mm = []
        if key_geom is not None:
            x_values_mm.extend([
                key_geom['x_left_cm'] * 10.0,
                key_geom['x_right_cm'] * 10.0,
            ])
        else:
            if el.get('x_left_cm') is not None:
                x_values_mm.append(_to_float(el.get('x_left_cm'), 0.0) * 10.0)
            if el.get('x_right_cm') is not None:
                x_values_mm.append(_to_float(el.get('x_right_cm'), interior_x_max_mm / 10.0) * 10.0)

        for x_raw in x_values_mm:
            # Ignorer les bords externes de l'intérieur utile.
            if x_raw <= 1.0 or x_raw >= interior_x_max_mm - 1.0:
                continue
            left_x = _normalize_web_separator_left_x(x_raw)
            edge_tol = max(2.0, inferred_t_mm * 0.5)
            # Exclure les candidats collés aux montants latéraux du caisson.
            if left_x <= t_lr + edge_tol:
                continue
            if left_x >= (L_mm - t_lr - inferred_t_mm - edge_tol):
                continue
            boundary_candidates_left_x.append(left_x)

    # Dédoublonnage tolérant des candidats.
    candidate_merge_tol_mm = max(2.0, inferred_t_mm + 2.0)
    unique_candidates = []
    for x_left in sorted(boundary_candidates_left_x):
        if not unique_candidates or abs(unique_candidates[-1] - x_left) > candidate_merge_tol_mm:
            unique_candidates.append(x_left)

    for x_left in unique_candidates:
        sep_center_x = x_left + inferred_t_mm / 2.0
        existing_divider = None
        for div in cabinet['vertical_dividers']:
            if abs(float(div.get('position_x', 0.0)) - sep_center_x) <= candidate_merge_tol_mm:
                existing_divider = div
                break

        if existing_divider is None:
            cabinet['vertical_dividers'].append({
                'position_x': sep_center_x,
                'thickness': inferred_t_mm,
                'material': mat_name,
                'bottom_y': 0.0,
                'top_y': interior_h_mm,
                '_force_full_height': True,
                '_imported_separator': True,
                '_inferred_from_bounds': True,
            })
        else:
            # Si déjà présent, garantir qu'il reste visible sur toute la hauteur
            # dans les JSON web legacy où seules les limites X sont fiables.
            existing_divider['bottom_y'] = 0.0
            existing_divider['top_y'] = interior_h_mm
            existing_divider['_force_full_height'] = True
            existing_divider['_imported_separator'] = True

    divider_match_tol_mm = max(2.0, inferred_t_mm * 0.5)

    def _should_expand_divider_to_double(divider: dict) -> bool:
        div_center = float(divider.get('position_x', 0.0))
        div_thickness = max(1.0, float(divider.get('thickness', inferred_t_mm)))
        div_left = div_center - div_thickness / 2.0
        div_right = div_center + div_thickness / 2.0

        left_drawers = []
        right_drawers = []
        for pending in pending_drawers:
            x_left = float(pending.get('x_left_mm', 0.0))
            x_right = float(pending.get('x_right_mm', 0.0))
            y_bottom = float(pending.get('y_bottom_mm', 0.0))
            y_top = y_bottom + float(pending.get('height_mm', 0.0))
            if x_left < div_center and abs(x_right - div_center) <= div_thickness + divider_match_tol_mm:
                left_drawers.append((y_bottom, y_top))
            if x_right > div_center and abs(x_left - div_center) <= div_thickness + divider_match_tol_mm:
                right_drawers.append((y_bottom, y_top))

        for left_bottom, left_top in left_drawers:
            for right_bottom, right_top in right_drawers:
                overlap_y = min(left_top, right_top) - max(left_bottom, right_bottom)
                if overlap_y > divider_match_tol_mm:
                    return True
        return False

    expanded_dividers = []
    for divider in sorted(cabinet['vertical_dividers'], key=lambda d: float(d.get('position_x', 0.0))):
        if not _should_expand_divider_to_double(divider):
            expanded_dividers.append(divider)
            continue

        div_thickness = max(1.0, float(divider.get('thickness', inferred_t_mm)))
        div_center = float(divider.get('position_x', 0.0))
        left_divider = copy.deepcopy(divider)
        right_divider = copy.deepcopy(divider)
        left_divider['position_x'] = div_center - div_thickness / 2.0
        right_divider['position_x'] = div_center + div_thickness / 2.0
        left_divider['_imported_double_divider'] = True
        right_divider['_imported_double_divider'] = True
        expanded_dividers.extend([left_divider, right_divider])

    deduped_dividers = []
    final_divider_tol_mm = max(1.0, inferred_t_mm * 0.25)
    for divider in sorted(expanded_dividers, key=lambda d: float(d.get('position_x', 0.0))):
        divider_x = float(divider.get('position_x', 0.0))
        existing_divider = None
        for kept in deduped_dividers:
            if abs(float(kept.get('position_x', 0.0)) - divider_x) <= final_divider_tol_mm:
                existing_divider = kept
                break

        if existing_divider is None:
            deduped_dividers.append(divider)
            continue

        existing_divider['bottom_y'] = min(float(existing_divider.get('bottom_y', 0.0)), float(divider.get('bottom_y', 0.0)))
        existing_divider['top_y'] = max(float(existing_divider.get('top_y', 0.0)), float(divider.get('top_y', 0.0)))
        existing_divider['thickness'] = max(float(existing_divider.get('thickness', inferred_t_mm)), float(divider.get('thickness', inferred_t_mm)))
        existing_divider['_force_full_height'] = bool(existing_divider.get('_force_full_height', False) or divider.get('_force_full_height', False))
        existing_divider['_imported_separator'] = True
        if divider.get('_imported_double_divider'):
            existing_divider['_imported_double_divider'] = True

    cabinet['vertical_dividers'] = deduped_dividers

    cabinet['door_props'] = door_props
    has_native_import_elements = bool(cabinet['vertical_dividers'] or cabinet['shelves'] or pending_drawers or pending_doors)
    # Si le JSON contient deja des elements de structure, on force le rendu natif
    # (separations -> zones -> tiroirs) au lieu du rendu geometrique explicite.
    if has_native_import_elements:
        cabinet['web_render_geometry'] = None
    else:
        cabinet['web_render_geometry'] = _build_web_render_geometry(payload, meuble, inferred_t_mm)

    # IMPORTANT: ordre impose pour fiabiliser l'import web.
    # 1) Separations/etageres posees  2) zones calculees  3) tiroirs importes.
    try:
        zones_for_import = calculate_all_zones_2d(cabinet, include_all_elements=False)
    except Exception:
        zones_for_import = []

    tol_mm = max(2.0, inferred_t_mm * 0.5)

    for shelf in cabinet['shelves']:
        stored_x_start = shelf.get('stored_shelf_x_start_mm')
        stored_width = shelf.get('stored_shelf_width_mm')
        if stored_x_start is None or stored_width is None or not zones_for_import:
            shelf['zone_id'] = None
            continue

        shelf_x_left = float(stored_x_start)
        shelf_x_right = shelf_x_left + float(stored_width)

        chosen_zone = None
        best_score = 0.0
        for zone in zones_for_import:
            zx_min = float(zone['x_min'])
            zx_max = float(zone['x_max'])
            x_match = abs(zx_min - shelf_x_left) <= tol_mm and abs(zx_max - shelf_x_right) <= tol_mm
            if x_match:
                chosen_zone = zone
                break
            overlap_x = max(0.0, min(shelf_x_right, zx_max) - max(shelf_x_left, zx_min))
            if overlap_x <= 0.0:
                continue
            if overlap_x > best_score + 1e-6:
                best_score = overlap_x
                chosen_zone = zone

        if chosen_zone is not None:
            zone_x_min = float(chosen_zone['x_min'])
            zone_x_max = float(chosen_zone['x_max'])
            shelf['zone_id'] = chosen_zone['id']
            shelf['stored_zone_coords'] = {
                'x_min': chosen_zone['x_min'],
                'x_max': chosen_zone['x_max'],
                'y_min': chosen_zone['y_min'],
                'y_max': chosen_zone['y_max'],
            }
            shelf['stored_shelf_x_start_mm'] = max(shelf_x_left, zone_x_min)
            shelf['stored_shelf_width_mm'] = max(0.0, min(shelf_x_right, zone_x_max) - shelf['stored_shelf_x_start_mm'])
            if shelf['stored_shelf_width_mm'] <= 0.0:
                shelf['stored_shelf_x_start_mm'] = zone_x_min
                shelf['stored_shelf_width_mm'] = max(0.0, zone_x_max - zone_x_min)
        else:
            shelf['zone_id'] = None

    def _best_zone_for_rect(x_left: float, x_right: float, y_bottom: float, y_top: float):
        exact_zone = None
        for zone in zones_for_import:
            x_match = abs(float(zone['x_min']) - x_left) <= tol_mm and abs(float(zone['x_max']) - x_right) <= tol_mm
            y_contains = y_bottom >= float(zone['y_min']) - tol_mm and y_top <= float(zone['y_max']) + tol_mm
            if x_match and y_contains:
                exact_zone = zone
                break
        if exact_zone is not None:
            return exact_zone

        best_zone = None
        best_score = 0.0
        for zone in zones_for_import:
            zx_min = float(zone['x_min'])
            zx_max = float(zone['x_max'])
            zy_min = float(zone['y_min'])
            zy_max = float(zone['y_max'])
            overlap_x = max(0.0, min(x_right, zx_max) - max(x_left, zx_min))
            overlap_y = max(0.0, min(y_top, zy_max) - max(y_bottom, zy_min))
            if overlap_x <= 0.0 or overlap_y <= 0.0:
                continue
            score = (overlap_x * overlap_y) - (0.1 * (abs(y_bottom - zy_min) + abs(y_top - zy_max)))
            if score > best_score + 1e-6:
                best_score = score
                best_zone = zone
        return best_zone

    for pending in pending_drawers:
        dr = copy.deepcopy(drawer_default)
        dr['drawer_system'] = drawer_system_default
        dr['drawer_tech_type'] = drawer_tech_default
        dr['drawer_face_H_raw'] = float(pending['height_mm'])
        dr['drawer_bottom_offset'] = float(pending['y_bottom_mm'])
        dr['drawer_face_thickness'] = 19.0
        dr['drawer_gap'] = 2.0
        dr['zone_id'] = None
        dr['material'] = 'Matière Tiroir'
        dr['x_left_mm'] = float(pending['x_left_mm'])
        dr['x_right_mm'] = float(pending['x_right_mm'])
        dr['_use_explicit_x_bounds'] = True

        x_left = dr['x_left_mm']
        x_right = dr['x_right_mm']
        y_bottom = dr['drawer_bottom_offset']
        y_top = y_bottom + dr['drawer_face_H_raw']

        exact_zone = None
        for zone in zones_for_import:
            x_match = abs(float(zone['x_min']) - x_left) <= tol_mm and abs(float(zone['x_max']) - x_right) <= tol_mm
            y_contains = y_bottom >= float(zone['y_min']) - tol_mm and y_top <= float(zone['y_max']) + tol_mm
            if x_match and y_contains:
                exact_zone = zone
                break

        chosen_zone = exact_zone
        if chosen_zone is not None:
            dr['zone_id'] = chosen_zone['id']
            dr['stored_zone_coords'] = {
                'x_min': chosen_zone['x_min'],
                'x_max': chosen_zone['x_max'],
                'y_min': chosen_zone['y_min'],
                'y_max': chosen_zone['y_max'],
            }
        if chosen_zone is None:
            best_zone = None
            best_score = 0.0
            for zone in zones_for_import:
                zx_min = float(zone['x_min'])
                zx_max = float(zone['x_max'])
                zy_min = float(zone['y_min'])
                zy_max = float(zone['y_max'])

                overlap_x = max(0.0, min(x_right, zx_max) - max(x_left, zx_min))
                overlap_y = max(0.0, min(y_top, zy_max) - max(y_bottom, zy_min))
                if overlap_x <= 0.0 or overlap_y <= 0.0:
                    continue

                # Score mixte: surface de chevauchement + bonus alignement vertical proche.
                overlap_area = overlap_x * overlap_y
                y_penalty = abs(y_bottom - zy_min) + abs(y_top - zy_max)
                score = overlap_area - (0.1 * y_penalty)
                if score > best_score + 1e-6:
                    best_score = score
                    best_zone = zone
            if best_zone is not None:
                chosen_zone = best_zone
                dr['zone_id'] = chosen_zone['id']
                dr['stored_zone_coords'] = {
                    'x_min': chosen_zone['x_min'],
                    'x_max': chosen_zone['x_max'],
                    'y_min': chosen_zone['y_min'],
                    'y_max': chosen_zone['y_max'],
                }
                zone_x_min = float(chosen_zone['x_min'])
                zone_x_max = float(chosen_zone['x_max'])
                dr['x_left_mm'] = max(float(dr['x_left_mm']), zone_x_min)
                dr['x_right_mm'] = min(float(dr['x_right_mm']), zone_x_max)
                if dr['x_right_mm'] <= dr['x_left_mm']:
                    dr['x_left_mm'] = zone_x_min
                    dr['x_right_mm'] = zone_x_max
            else:
                dr['zone_id'] = None

        cabinet['drawers'].append(dr)

    for pending in pending_doors:
        door_entry = {
            'x_left_mm': float(pending['x_left_mm']),
            'x_right_mm': float(pending['x_right_mm']),
            'door_bottom_offset': float(pending['y_bottom_mm']),
            'door_face_H_raw': float(pending['height_mm']),
            'door_thickness': float(door_props.get('door_thickness', 19.0)),
            'door_gap': float(door_props.get('door_gap', 2.0)),
            'door_opening': pending.get('door_opening') if pending.get('door_opening') in {'left', 'right'} else None,
            'has_explicit_opening': bool(pending.get('has_explicit_opening', False)),
            'material': door_props.get('material', 'Matière Porte'),
            'zone_id': None,
        }
        if pending.get('is_double_leaf'):
            door_entry['is_double_leaf'] = True

        x_left = door_entry['x_left_mm']
        x_right = door_entry['x_right_mm']
        y_bottom = door_entry['door_bottom_offset']
        y_top = y_bottom + door_entry['door_face_H_raw']

        exact_zone = None
        for zone in zones_for_import:
            x_match = abs(float(zone['x_min']) - x_left) <= tol_mm and abs(float(zone['x_max']) - x_right) <= tol_mm
            y_contains = y_bottom >= float(zone['y_min']) - tol_mm and y_top <= float(zone['y_max']) + tol_mm
            if x_match and y_contains:
                exact_zone = zone
                break

        if exact_zone is not None:
            door_entry['zone_id'] = exact_zone['id']
            door_entry['stored_zone_coords'] = {
                'x_min': exact_zone['x_min'],
                'x_max': exact_zone['x_max'],
                'y_min': exact_zone['y_min'],
                'y_max': exact_zone['y_max'],
            }

        # Règle de fallback demandée (si JSON n'indique pas explicitement le sens):
        # 1) porte zone gauche => ouverture à droite
        # 2) porte zone droite => ouverture à gauche
        # 3) porte pleine largeur => double battant auto
        interior_left = t_lr
        interior_right = max(t_lr, L_mm - t_lr)
        interior_mid = (interior_left + interior_right) / 2.0
        full_width = (
            abs(door_entry['x_left_mm'] - interior_left) <= tol_mm
            and abs(door_entry['x_right_mm'] - interior_right) <= tol_mm
        )
        if full_width:
            door_entry['is_double_leaf'] = True
        elif not door_entry.get('has_explicit_opening', False):
            center_x = (door_entry['x_left_mm'] + door_entry['x_right_mm']) / 2.0
            if center_x <= interior_mid:
                door_entry['door_opening'] = 'left'
            else:
                door_entry['door_opening'] = 'right'

        if door_entry.get('door_opening') not in {'left', 'right'}:
            door_entry['door_opening'] = 'right'

        cabinet['imported_doors'].append(door_entry)

    if cabinet['imported_doors']:
        cabinet['door_props']['has_door'] = False

    cabinet['shelves'] = sorted(cabinet['shelves'], key=lambda s: float(s.get('height', 0.0)))
    cabinet['drawers'] = sorted(cabinet['drawers'], key=lambda d: float(d.get('drawer_bottom_offset', 0.0)))
    cabinet['imported_doors'] = sorted(
        cabinet['imported_doors'],
        key=lambda d: (float(d.get('door_bottom_offset', 0.0)), float(d.get('x_left_mm', 0.0)))
    )
    return cabinet


def load_web_config_json_state():
    if 'web_json_loader' not in st.session_state:
        return
    uploaded_file = st.session_state.web_json_loader
    if uploaded_file is None:
        return

    try:
        raw = uploaded_file.getvalue()
        payload = json.loads(raw.decode('utf-8-sig'))
        if not isinstance(payload, dict):
            st.error("Le fichier JSON doit contenir un objet racine.")
            return

        meuble = payload.get('meuble', {}) if isinstance(payload.get('meuble', {}), dict) else {}

        scene = _build_scene_from_web_config(payload)
        st.session_state['scene_cabinets'] = scene
        st.session_state['selected_cabinet_index'] = 0
        st.session_state['base_cabinet_index'] = 0

        project_type = str(payload.get('furniture_type') or payload.get('type') or 'Projet Web').strip() or 'Projet Web'
        st.session_state['project_name'] = f"{project_type} - Import JSON"
        st.session_state['panneau_decor'] = str(
            payload.get('mat')
            or meuble.get('matLabel')
            or meuble.get('mat')
            or st.session_state.get('panneau_decor', 'BLANC')
        )
        st.success(f"Projet JSON importé : {len(scene)} caisson(s) détecté(s) et chargés dans la scène.")
        st.rerun()
    except Exception as e:
        st.error(f"Erreur import JSON configurateur : {e}")

# Callbacks
def update_selected_cabinet_dim(key):
    cabinet = get_selected_cabinet()
    widget_key = f"{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state: cabinet['dims'][key] = st.session_state[widget_key]

def update_selected_cabinet_base_element(element_key):
    """Met à jour l'état d'un élément de base du caisson (fond, montants, traverses)."""
    cabinet = get_selected_cabinet()
    idx = st.session_state.selected_cabinet_index
    widget_key = f"base_element_{element_key}_{idx}"
    if cabinet and widget_key in st.session_state:
        if 'base_elements' not in cabinet:
            cabinet['base_elements'] = {
                'has_back_panel': True,
                'has_left_upright': True,
                'has_right_upright': True,
                'has_bottom_traverse': True,
                'has_top_traverse': True
            }
        cabinet['base_elements'][element_key] = st.session_state[widget_key]

def update_selected_cabinet_door(key):
    cabinet = get_selected_cabinet()
    widget_key = f"{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state:
        if 'door_props' not in cabinet: cabinet['door_props'] = get_default_door_props_19()
        cabinet['door_props'][key] = st.session_state[widget_key]
        if key == 'has_door' and st.session_state[widget_key] is True:
            if 'drawer_props' in cabinet: cabinet['drawer_props']['has_drawer'] = False

def update_selected_cabinet_drawer(key):
    # Ancien système (compatibilité) - à supprimer progressivement
    cabinet = get_selected_cabinet()
    # Pour zone_id, la clé du widget est "drawer_zone_{idx}" et non "zone_id_{idx}"
    if key == 'zone_id':
        widget_key = f"drawer_zone_{st.session_state.selected_cabinet_index}"
    else:
        widget_key = f"drawer_{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state:
        if 'drawer_props' not in cabinet: cabinet['drawer_props'] = get_default_drawer_props_19()
        cabinet['drawer_props'][key] = st.session_state[widget_key]
        if key == 'has_drawer' and st.session_state[widget_key] is True:
            if 'door_props' in cabinet: cabinet['door_props']['has_door'] = False

def add_drawer_callback():
    """Ajoute un nouveau tiroir en mode preview (pose en 2 temps)."""
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'drawers' not in cabinet: cabinet['drawers'] = []
        # Pose en 2 temps: on crée un "pending" au lieu d'ajouter directement
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        st.session_state['pending_placement'] = {
            'kind': 'drawer',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': get_default_drawer_props()
        }

def add_drawers_stack_callback():
    """Ajoute plusieurs tiroirs empilés (pose en 2 temps), sans demander de dimensions à l'utilisateur."""
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'drawers' not in cabinet: cabinet['drawers'] = []
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        p = get_default_drawer_props()
        # Mode empilement : dimensions (hauteur/offset) seront calculées automatiquement à partir de la zone
        p['stack_count'] = 3
        p['_stack_mode'] = True
        st.session_state['pending_placement'] = {
            'kind': 'drawer_stack',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': p
        }

def update_drawer_prop(drawer_index, key):
    """Met à jour une propriété d'un tiroir existant."""
    cabinet = get_selected_cabinet()
    if key == 'drawer_tech_type': widget_key = f"drawer_tech_type_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_system': widget_key = f"drawer_system_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_handle_type': widget_key = f"drawer_handle_type_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'zone_id': widget_key = f"drawer_zone_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_bottom_offset': widget_key = f"drawer_bottom_offset_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_face_H_raw': widget_key = f"drawer_face_H_raw_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_face_thickness': widget_key = f"drawer_face_thickness_{st.session_state.selected_cabinet_index}_{drawer_index}"
    elif key == 'drawer_gap': widget_key = f"drawer_gap_{st.session_state.selected_cabinet_index}_{drawer_index}"
    else: widget_key = f"drawer_{key}_{st.session_state.selected_cabinet_index}_{drawer_index}"
    if cabinet and widget_key in st.session_state:
        if 'drawers' in cabinet and drawer_index < len(cabinet['drawers']): 
            cabinet['drawers'][drawer_index][key] = st.session_state[widget_key]

def delete_drawer_callback(drawer_index):
    """Supprime un tiroir."""
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'drawers' in cabinet and drawer_index < len(cabinet['drawers']):
            cabinet['drawers'].pop(drawer_index)
            st.rerun()

def update_drawer_material(drawer_index):
    """Met à jour la matière d'un tiroir."""
    widget_key = f"drawer_material_{st.session_state.selected_cabinet_index}_{drawer_index}"
    cabinet = get_selected_cabinet()
    if cabinet and widget_key in st.session_state:
        if 'drawers' in cabinet and drawer_index < len(cabinet['drawers']): 
            cabinet['drawers'][drawer_index]['material'] = st.session_state[widget_key]

def add_shelf_callback():
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'shelves' not in cabinet: cabinet['shelves'] = []
        # Pose en 2 temps: on crée un "pending" au lieu d'ajouter directement
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        st.session_state['pending_placement'] = {
            'kind': 'shelf',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': get_default_shelf_props()
        }


def add_fixed_shelves_stack_callback():
    """Ajoute un assemblage d'etageres fixes (pose en 2 temps)."""
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'shelves' not in cabinet:
            cabinet['shelves'] = []
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return

        p = get_default_shelf_props()
        p['shelf_type'] = 'fixe'
        p['stack_count'] = 3
        p['_stack_mode'] = True
        st.session_state['pending_placement'] = {
            'kind': 'shelf_stack',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': p,
        }

def update_shelf_prop(shelf_index, key):
    cabinet = get_selected_cabinet()
    if key == 'shelf_type': widget_key = f"shelf_t_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'height': widget_key = f"shelf_h_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'thickness': widget_key = f"shelf_e_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'mobile_machining_type': widget_key = f"shelf_m_type_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'custom_holes_above': widget_key = f"shelf_c_above_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'custom_holes_below': widget_key = f"shelf_c_below_{st.session_state.selected_cabinet_index}_{shelf_index}"
    elif key == 'zone_id': widget_key = f"shelf_zone_{st.session_state.selected_cabinet_index}_{shelf_index}"
    else: widget_key = f"shelf_{key[0]}_{st.session_state.selected_cabinet_index}_{shelf_index}"
    if cabinet and widget_key in st.session_state:
        if 'shelves' in cabinet and shelf_index < len(cabinet['shelves']): 
            cabinet['shelves'][shelf_index][key] = st.session_state[widget_key]

def delete_shelf_callback(shelf_index):
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'shelves' in cabinet and shelf_index < len(cabinet['shelves']):
            cabinet['shelves'].pop(shelf_index)
            st.rerun()

def update_selected_cabinet_material(key):
    cabinet = get_selected_cabinet()
    widget_key = f"{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state: cabinet[key] = st.session_state[widget_key]
def update_selected_cabinet_door_material(key):
    cabinet = get_selected_cabinet()
    widget_key = f"door_{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state: cabinet['door_props']['material'] = st.session_state[widget_key]

def update_hinge_count(cabinet_index):
    """Met à jour le nombre de charnières personnalisées."""
    if cabinet_index < len(st.session_state['scene_cabinets']):
        cabinet = st.session_state['scene_cabinets'][cabinet_index]
        if 'door_props' in cabinet:
            widget_key = f"num_hinges_{cabinet_index}"
            if widget_key in st.session_state:
                num_hinges = int(st.session_state[widget_key])
                custom_positions = cabinet['door_props'].get('custom_hinge_positions', [])
                door_height = cabinet['dims']['H_raw']
                
                # Ajuster la liste selon le nouveau nombre
                if len(custom_positions) < num_hinges:
                    # Ajouter des positions par défaut
                    for i in range(len(custom_positions), num_hinges):
                        pos = (i + 1) * door_height / (num_hinges + 1)
                        custom_positions.append(pos)
                else:
                    # Retirer les positions en trop
                    custom_positions = custom_positions[:num_hinges]
                
                cabinet['door_props']['custom_hinge_positions'] = custom_positions

def update_hinge_position(cabinet_index, hinge_index):
    """Met à jour la position d'une charnière personnalisée."""
    if cabinet_index < len(st.session_state['scene_cabinets']):
        cabinet = st.session_state['scene_cabinets'][cabinet_index]
        if 'door_props' in cabinet:
            widget_key = f"hinge_pos_{cabinet_index}_{hinge_index}"
            if widget_key in st.session_state:
                if 'custom_hinge_positions' not in cabinet['door_props']:
                    cabinet['door_props']['custom_hinge_positions'] = []
                custom_positions = cabinet['door_props']['custom_hinge_positions']
                
                # S'assurer que la liste est assez longue
                while len(custom_positions) <= hinge_index:
                    custom_positions.append(0.0)
                
                custom_positions[hinge_index] = float(st.session_state[widget_key])
                cabinet['door_props']['custom_hinge_positions'] = custom_positions
def update_selected_cabinet_drawer_material(key):
    cabinet = get_selected_cabinet()
    widget_key = f"drawer_{key}_{st.session_state.selected_cabinet_index}"
    if cabinet and widget_key in st.session_state: cabinet['drawer_props']['material'] = st.session_state[widget_key]
def update_shelf_material(shelf_index, key):
    widget_key = f"shelf_m_{st.session_state.selected_cabinet_index}_{shelf_index}"
    cabinet = get_selected_cabinet()
    if cabinet and widget_key in st.session_state:
        if 'shelves' in cabinet and shelf_index < len(cabinet['shelves']): cabinet['shelves'][shelf_index]['material'] = st.session_state[widget_key]

def add_cabinet(origin_type='central'):
    if origin_type == 'central':
        if st.session_state['scene_cabinets']: return
        new_cabinet = {
            'dims': get_default_dims_19(), 'debit_data': get_default_debit_data(), 'name': "Caisson 0 (Central)",
            'parent_index': None, 'attachment_dir': None, 'door_props': get_default_door_props_19(),
            'drawer_props': get_default_drawer_props_19(), 'drawers': [], 'shelves': [], 'material_body': 'Matière Corps',
            'vertical_dividers': [],  # Nouveaux montants verticaux secondaires
            'vertical_shelves': [],  # Étagères verticales
            'joues': {
                'gauche': get_default_joue_props(),
                'droite': get_default_joue_props(),
                'dessus': get_default_joue_props(),
                'dessous': get_default_joue_props(),
            }
        }
        st.session_state['scene_cabinets'].append(new_cabinet)
        st.session_state['selected_cabinet_index'] = 0
        st.session_state['base_cabinet_index'] = 0
    else: 
        base_index = st.session_state.get('base_cabinet_index', 0)
        if base_index is None or base_index >= len(st.session_state['scene_cabinets']):
            st.error("Aucun caisson de base sélectionné.")
            return
        base_caisson = st.session_state['scene_cabinets'][base_index]
        new_cabinet = {
            'dims': copy.deepcopy(base_caisson['dims']), 'debit_data': get_default_debit_data(),
            'parent_index': base_index, 'attachment_dir': origin_type, 'door_props': get_default_door_props_19(),
            'drawer_props': get_default_drawer_props_19(), 'drawers': [], 'shelves': [], 'material_body': 'Matière Corps',
            'vertical_dividers': [],  # Nouveaux montants verticaux secondaires
            'vertical_shelves': [],  # Étagères verticales
            'joues': {
                'gauche': get_default_joue_props(),
                'droite': get_default_joue_props(),
                'dessus': get_default_joue_props(),
                'dessous': get_default_joue_props(),
            }
        }
        if origin_type == 'right': new_name = f"D de {base_index}"
        elif origin_type == 'left': new_name = f"G de {base_index}"
        else: new_name = f"H de {base_index}"
        new_cabinet['name'] = f"Caisson {len(st.session_state['scene_cabinets'])} ({new_name})"
        st.session_state['scene_cabinets'].append(new_cabinet)
        new_index = len(st.session_state['scene_cabinets']) - 1
        st.session_state['selected_cabinet_index'] = new_index
        st.session_state['base_cabinet_index'] = st.session_state['selected_cabinet_index']

def clear_scene():
    st.session_state['scene_cabinets'] = []
    st.session_state['selected_cabinet_index'] = None
    st.session_state['base_cabinet_index'] = 0

def delete_selected_cabinet():
    idx = st.session_state.get('selected_cabinet_index')
    if idx is None or idx >= len(st.session_state['scene_cabinets']): return
    indices_to_remove = set()
    queue = [idx]
    while queue:
        curr = queue.pop()
        if curr not in indices_to_remove:
            indices_to_remove.add(curr)
            for i, c in enumerate(st.session_state['scene_cabinets']):
                if c['parent_index'] == curr: queue.append(i)
    new_scene = []
    map_old_new = {}
    counter = 0
    for i, c in enumerate(st.session_state['scene_cabinets']):
        if i not in indices_to_remove:
            map_old_new[i] = counter
            new_scene.append(c)
            counter += 1
    for c in new_scene:
        if c['parent_index'] is not None: c['parent_index'] = map_old_new.get(c['parent_index'], None) 
    st.session_state['scene_cabinets'] = new_scene
    st.session_state['selected_cabinet_index'] = 0 if new_scene else None
    st.session_state['base_cabinet_index'] = 0

# Callbacks pour les montants verticaux secondaires
def add_vertical_divider_callback():
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'vertical_dividers' not in cabinet:
            cabinet['vertical_dividers'] = []
        # Pose en 2 temps: on crée un "pending" au lieu d'ajouter directement
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        new_divider = get_default_vertical_divider_props()
        # S'assurer que la position est valide (entre les montants)
        dims = cabinet['dims']
        t_lr = dims['t_lr_raw']
        L_raw = dims['L_raw']
        # Position par défaut au milieu si c'est le premier montant
        if len(cabinet['vertical_dividers']) == 0:
            new_divider['position_x'] = (L_raw - 2*t_lr) / 2.0 + t_lr
            new_divider['zone_id'] = None  # Premier montant : pas de zone assignée, placement libre dans Zone 0
        else:
            # Positionner après le dernier montant
            last_div = max(cabinet['vertical_dividers'], key=lambda d: d['position_x'])
            new_divider['position_x'] = min(last_div['position_x'] + 200.0, L_raw - t_lr - 50.0)
            new_divider['zone_id'] = None  # Sera assigné par l'utilisateur via le selectbox
        st.session_state['pending_placement'] = {
            'kind': 'vertical_divider',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': new_divider
        }

def add_vertical_divider_double_callback():
    """Ajoute un double montant secondaire (2 montants côte à côte, sans jeu) en mode preview."""
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'vertical_dividers' not in cabinet:
            cabinet['vertical_dividers'] = []
        # Pose en 2 temps : on crée un pending spécifique
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        new_divider = get_default_vertical_divider_props()
        dims = cabinet['dims']
        t_lr = dims['t_lr_raw']
        L_raw = dims['L_raw']
        # Centre par défaut au milieu du caisson
        new_divider['position_x'] = (L_raw - 2 * t_lr) / 2.0 + t_lr
        new_divider['zone_id'] = None
        new_divider['double'] = True  # marqueur pour l'UI
        st.session_state['pending_placement'] = {
            'kind': 'vertical_divider_double',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': new_divider
        }

def update_vertical_divider_prop(divider_index, key):
    cabinet = get_selected_cabinet()
    if key == 'zone_id':
        widget_key = f"divider_zone_{st.session_state.selected_cabinet_index}_{divider_index}"
    else:
        widget_key = f"divider_{key}_{st.session_state.selected_cabinet_index}_{divider_index}"
    if cabinet and widget_key in st.session_state:
        if 'vertical_dividers' in cabinet and divider_index < len(cabinet['vertical_dividers']):
            cabinet['vertical_dividers'][divider_index][key] = st.session_state[widget_key]

def delete_vertical_divider_callback(divider_index):
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'vertical_dividers' in cabinet and divider_index < len(cabinet['vertical_dividers']):
            cabinet['vertical_dividers'].pop(divider_index)
            st.rerun()

def update_vertical_divider_material(divider_index):
    cabinet = get_selected_cabinet()
    widget_key = f"divider_material_{st.session_state.selected_cabinet_index}_{divider_index}"
    if cabinet and widget_key in st.session_state:
        if 'vertical_dividers' in cabinet and divider_index < len(cabinet['vertical_dividers']):
            cabinet['vertical_dividers'][divider_index]['material'] = st.session_state[widget_key]

# Callbacks pour les étagères verticales
def add_vertical_shelf_callback():
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'vertical_shelves' not in cabinet:
            cabinet['vertical_shelves'] = []
        # Pose en 2 temps: on crée un "pending" au lieu d'ajouter directement
        if st.session_state.get('pending_placement') is not None:
            st.warning("Une pose est déjà en cours. Validez ou annulez avant d'ajouter un nouvel élément.")
            return
        new_shelf = get_default_vertical_shelf_props()
        # S'assurer que la position est valide
        dims = cabinet['dims']
        t_lr = dims['t_lr_raw']
        L_raw = dims['L_raw']
        H_raw = dims['H_raw']
        # Position par défaut au milieu si c'est la première étagère
        if len(cabinet['vertical_shelves']) == 0:
            new_shelf['position_x'] = (L_raw - 2*t_lr) / 2.0 + t_lr
            new_shelf['bottom_y'] = 100.0
            new_shelf['top_y'] = min(400.0, H_raw - 100.0)
        else:
            # Positionner après la dernière étagère
            last_shelf = max(cabinet['vertical_shelves'], key=lambda s: s['position_x'])
            new_shelf['position_x'] = min(last_shelf['position_x'] + 200.0, L_raw - t_lr - 50.0)
            new_shelf['bottom_y'] = 100.0
            new_shelf['top_y'] = min(400.0, H_raw - 100.0)
        st.session_state['pending_placement'] = {
            'kind': 'vertical_shelf',
            'cabinet_index': st.session_state.get('selected_cabinet_index'),
            'props': new_shelf
        }

def update_vertical_shelf_prop(shelf_index, key):
    cabinet = get_selected_cabinet()
    if key == 'zone_id':
        widget_key = f"vertical_shelf_zone_{st.session_state.selected_cabinet_index}_{shelf_index}"
    else:
        widget_key = f"vertical_shelf_{key}_{st.session_state.selected_cabinet_index}_{shelf_index}"
    if cabinet and widget_key in st.session_state:
        if 'vertical_shelves' in cabinet and shelf_index < len(cabinet['vertical_shelves']):
            cabinet['vertical_shelves'][shelf_index][key] = st.session_state[widget_key]

def delete_vertical_shelf_callback(shelf_index):
    cabinet = get_selected_cabinet()
    if cabinet:
        if 'vertical_shelves' in cabinet and shelf_index < len(cabinet['vertical_shelves']):
            cabinet['vertical_shelves'].pop(shelf_index)
            st.rerun()

def update_vertical_shelf_material(shelf_index):
    cabinet = get_selected_cabinet()
    widget_key = f"vertical_shelf_material_{st.session_state.selected_cabinet_index}_{shelf_index}"
    if cabinet and widget_key in st.session_state:
        if 'vertical_shelves' in cabinet and shelf_index < len(cabinet['vertical_shelves']):
            cabinet['vertical_shelves'][shelf_index]['material'] = st.session_state[widget_key]